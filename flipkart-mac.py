from selenium import webdriver
import time
from openpyxl import Workbook
from selenium.webdriver.common.keys import Keys
import os
import pandas as pd 

driver = webdriver.Firefox()


#excel
my_wb = Workbook()
my_sheet = my_wb.active

#driver navigate
driver.get("https://www.flipkart.com/search?q=laptops&as=on&as-show=on&otracker=AS_Query_TrendingAutoSuggest_0_0_na_na_na&otracker1=AS_Query_TrendingAutoSuggest_0_0_na_na_na&as-pos=0&as-type=HISTORY&suggestionId=laptops&requestId=3fb848a6-5364-4ab9-a213-70c0f402acb8")
driver.maximize_window()

#url
url = driver.current_url
print(url)


title = driver.title[0:7]
html = driver.find_element_by_tag_name('html')
productname = driver.find_elements_by_class_name("_2cLu-l")
price_product = driver.find_elements_by_class_name("_1uv9Cb")



product_len = len(productname)
print(product_len)

product_price = len(price_product)
print(product_price)
for u in range(100):
     html.send_keys(Keys.PAGE_DOWN)
for  k in range(100):
      html.send_keys(Keys.PAGE_UP)


for i , data in enumerate(productname):
    for j , price in enumerate(price_product):
        c2 = my_sheet.cell(row = j+1 , column=3)
        c2.value = price.text[0:7]
    c1 = my_sheet.cell(row=i+1 , column=2)
    c1.value = data.text
    c0 = my_sheet.cell(row = i+1, column = 1)
    c0.value = i+1
    print('Process Done {}'.format(i))
    my_wb.save("page1-{}.xlsx".format(title))
os.system("open page1-{}.xlsx".format(title))
time.sleep(5)
driver.maximize_window()
#page2 process
driver.find_element_by_xpath('//a[@class="_3fVaIS"]').click()
time.sleep(5)
time.sleep(8)


#url
print(driver.current_url)
title = driver.title[0:7]
driver.maximize_window()
#product and price
productname1 = driver.find_elements_by_class_name("_2cLu-l")
price_product1 = driver.find_elements_by_class_name("_1uv9Cb")



for i , data in enumerate(productname1):
    for j , price in enumerate(price_product1):
        c2 = my_sheet.cell(row = j+1 , column=3)
        c2.value = price.text[0:7]
    c1 = my_sheet.cell(row=i+1 , column=2)
    c1.value = data.text
    c0 = my_sheet.cell(row = i+1, column = 1)
    c0.value = i+1
    print('Process Done {}'.format(i))
    my_wb.save("page2-{}.xlsx".format(title))
    my_wb.save("page2-{}.csv".format(title))

#flipkart-page3
driver.find_element_by_xpath('//a[@class="_3fVaIS"]').click()
driver.minimize_window()
os.system("open page2-{}.xlsx".format(title))
driver.get("http://nahsolutions.in")
time.sleep(2)
driver.maximize_window()
html = driver.find_element_by_tag_name('html')
for i in range(40):
    html.send_keys(Keys.PAGE_DOWN)
for j in range(40):
    html.send_keys(Keys.PAGE_UP)    
time.sleep(2)    
driver.get("https://www.nahsolutions.in/about.html") 
html = driver.find_element_by_tag_name('html')
for i in range(40):
    html.send_keys(Keys.PAGE_DOWN)
for j in range(40):
    html.send_keys(Keys.PAGE_UP)         
time.sleep(5)
driver.get("http://google.com")
search = driver.find_element_by_name('q')
search.send_keys("Thank YOU.....")
time.sleep(20)
driver.maximize_window()
driver.close()

#read Excel file
data = pd.read_excel("page1-{}.xlsx".format(title))
print(data)